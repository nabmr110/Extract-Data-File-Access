import sys
import pandas as pd
import os
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Utility: Find nearest date above given row ===
def find_nearest_date(df, start_idx, date_col=0):

    max_rows = len(df)
    offset = 1

    while start_idx - offset >= 0 or start_idx + offset < max_rows:
        if start_idx - offset >= 0:
            # up_val = df.iloc[start_idx - offset, date_col]
            row = df.iloc[start_idx - offset]
            for val in row[:5]:
                if pd.notna(val) and isinstance(val, (datetime.date, pd.Timestamp)):
                    return pd.to_datetime(val)

        if start_idx + offset < max_rows:
            # down_val = df.iloc[start_idx + offset, date_col]
            row = df.iloc[start_idx + offset]
            for val in row[:5]:
                if pd.notna(val) and isinstance(val, (datetime.date, pd.Timestamp)):
                    return pd.to_datetime(val)
            
        offset += 1
    return None

# Get file path from argument
if len(sys.argv) < 2:
    print("No file path provided.", file=sys.stderr)
    sys.exit(1)

file_path = sys.argv[1]
ext = os.path.splitext(file_path)[1].lower()

try:
    # === Step 2: Load Excel File ===
    if ext in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path, sheet_name='Manpower hours', header=None)
        df = df.dropna(axis=1, how='all')  # Drop empty columns
        df = df.dropna(how='all') # Drop empty rows
        df = df[~(df.astype(str).apply(lambda row: row.str.strip().eq('').all(), axis=1))]
        df = df.reset_index(drop=True)


        # Debug: Save raw dataframe to CSV
        df.to_csv("debug_raw.csv", index=False)

        # Step 3: Extract Header Row
        header_row_idx = 0
        if len(df) <= header_row_idx:
            raise ValueError(f"Excel file does not have enough rows to extract header at index {header_row_idx}")

        df = df.iloc[header_row_idx:].copy()

        if df.empty or len(df) < 2:
            raise ValueError("Not enough rows to extract header and data")
        
        new_headers = df.iloc[header_row_idx]
        df = df[header_row_idx + 1:].copy()
        df.columns = new_headers
        df = df.reset_index(drop=True)

        df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], errors='coerce')
        print(f"Original raw rows: {len(df)}")

        # Step 4: Find CPBU Total Row
        cpbu_total_row = df[df.iloc[:, 1].astype(str).str.contains('CPBU Total', case=False, na=False)]
       
        # Step 6: Extract columns BQ to BU (Index 68 to 72)
        col_indices = list(range(68, 73))
        col_names = ['TOTAL HC', 'ATTN HC', 'NORMAL HRS', 'OT HRS', 'TOTAL HRS']

        summary_data = []

        if not cpbu_total_row.empty:
            for idx in cpbu_total_row.index:
                row = df.loc[idx]
                values = row.iloc[col_indices].values
                if not pd.isnull(values).all():
                    cpbu_date = find_nearest_date(df, idx)
                    print(f"Row Index: {idx}, Found Date: {cpbu_date}")
                    data_row = pd.DataFrame([values], columns=col_names)
                    data_row.insert(0, 'Date', cpbu_date)
                    data_row.insert(1, 'Section', 'P7 CPBU')
                    summary_data.append(data_row)

        # Step 5: Find QA-CPBU Sub-department Rows
        sub_dept_keywords = ['QA-BPA', 'QA-DREAME', 'QA-NANOJET', 'QA-WEBER', 'SQE/IQC']
        used_indices = set()  # Track globally used rows
        picked_dates = {}  # Track first occurrence per date for each keyword

        for keyword in sub_dept_keywords:
            picked_dates[keyword] = set()  # Initialize tracking for this keyword
            
            for idx, row in df.iterrows():
                if idx in used_indices:  
                    continue  # Skip already used rows

                if any(keyword in str(cell) for cell in row):
                    values = row.iloc[col_indices].values
                    
                    if not pd.isnull(values).all():  # Ensure meaningful values
                        row_date = find_nearest_date(df, idx)
                        
                        if row_date and row_date not in picked_dates[keyword]:  # Check if the date was already used
                            used_indices.add(idx)  # Mark row as used globally
                            picked_dates[keyword].add(row_date)  # Track the first occurrence for this date
                            print(f"QA-CPBU Index: {idx}, Found Date: {row_date}")

                            data_row = pd.DataFrame([values], columns=col_names)
                            data_row.insert(0, 'Date', row_date)
                            data_row.insert(1, 'Section', keyword)
                            summary_data.append(data_row)
                    

        # Step 7: Extract Grand Total under "P10A" section (columns F to J)

        # Search for the row where any of columns F to J contain "P10A"
        p10a_rows = df[df.iloc[:, 5].astype(str).str.contains('P10A', case=False, na=False)]
        
        if not p10a_rows.empty:
            p10a_index = p10a_rows.index[0]

            # Search for "Grand Total" below the P10A row, in column F (index 5)
            grand_total_row = df.iloc[p10a_index + 1:]
            grand_total_row = grand_total_row[grand_total_row.apply(
            lambda row: row.astype(str).str.contains('Grand Total', case=False, na=False).any(), axis=1)]


            if not grand_total_row.empty:
                col_names_p10a = ['TOTAL HC', 'ATTN HC', 'NORMAL HRS', 'OT HRS', 'TOTAL HRS']
                
                for idx, row in grand_total_row.iterrows():
                    data_row = pd.DataFrame([row.iloc[5:10].values], columns=col_names_p10a)
                    grand_total_date = row.iloc[0]
                    data_row.insert(0, 'Date', grand_total_date)
                    data_row.insert(1, 'Section', 'P10A CPBU')
                    print("Found Grand Total row under P10A:", row.iloc[5:10].values)
                    summary_data.append(data_row)


        if not summary_data:
            print("No matching data found for P7CPBU or QA-CPBU sub-departments.")
            sys.exit(0)

        final_df = pd.concat(summary_data, ignore_index=True)

        # Insert Grand Total rows per date ===
        grouped = final_df.groupby("Date")
        final_with_totals = []

        for date, group in grouped:
            final_with_totals.append(group.reset_index(drop=True))
            total_row = pd.DataFrame([{
                'Date': date,
                'Section': 'Grand Total',
                'TOTAL HC': group['TOTAL HC'].sum(),
                'ATTN HC': group['ATTN HC'].sum(),
                'NORMAL HRS': round(group['NORMAL HRS'].sum(), 2),
                'OT HRS': round(group['OT HRS'].sum(), 2),
                'TOTAL HRS': round(group['TOTAL HRS'].sum(), 2)
            }])
            final_with_totals.append(total_row)

        final_df = pd.concat(final_with_totals, ignore_index=True).reset_index(drop=True)

        if final_df.empty:
         print("No data to write - final_df is empty!", file=sys.stderr)
         sys.exit(1)

    elif ext == '.csv':
        print("CSV format not supported for this script.", file=sys.stderr)
        sys.exit(1)
    else:
        raise ValueError("Unsupported file format: " + ext)

    # === Step 8: Save processed file ===
    output_dir = os.path.join(os.path.dirname(os.path.dirname(file_path)), 'processed')
    os.makedirs(output_dir, exist_ok=True)

    output_filename = os.path.splitext(os.path.basename(file_path))[0] + '_summary.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    final_df.to_excel(output_path, index=False)

    # === Step 9: Excel Formatting ===
    processed_wb = load_workbook(output_path)
    processed_ws = processed_wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_idx, column_cells in enumerate(processed_ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row_idx, cell in enumerate(column_cells, 1):
            if cell.value is not None:
                if isinstance(cell.value, (datetime.datetime, datetime.date, pd.Timestamp)):
                    cell.number_format = 'yyyy-mm-dd'
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            if str(cell.value).strip().lower() == 'grand total':
                for target_cell in processed_ws[row_idx]:
                    target_cell.fill = yellow_fill

        processed_ws.column_dimensions[col_letter].width = max_length + 2
    
    # tempt_output_path = output_path + ".tmp"
    processed_wb.save(output_path)
    # os.replace(tempt_output_path, output_path)

    print(f"Found {len(df)} rows of data.")
    print("Output saved to:", output_path)

except Exception as e:
    print(f"Processing failed: {e}", file=sys.stderr)
    sys.exit(1)
